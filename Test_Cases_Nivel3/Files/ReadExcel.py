import openpyxl

archivo = openpyxl.load_workbook("C:\\Users\\Melany\\Documents\\Robot-Framework\\Test_Cases_Nivel3\\Files\\Libro1.xlsx")

print(archivo.sheetnames)

print("Página Activa " +archivo.active.title)

ac=archivo["Hoja1"]
print(ac['A2'].value)
print(ac['B2'].value)
print(ac['C2'].value)

sel=ac.cell(3,3)
print(sel.value)






