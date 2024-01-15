import openpyxl

archivo = openpyxl.load_workbook("C:\\Users\\Melany\\Documents\\Robot-Framework\\Test_Cases_Nivel3\\Files\\Libro1.xlsx")

def Numero_filas(hoja):
    ac = archivo[hoja]
    return ac.max_row


def Dato_columna(hoja, fila, col):
    ac = archivo[hoja]
    col = ac.cell(int(fila), int(col))
    return col.value

print(Numero_filas('Hoja2'))
print(Dato_columna('Hoja2', 1, 2))















