import openpyxl

archivo = openpyxl.load_workbook("C:\\Users\\Melany\\Documents\\Robot-Framework\\Test_Cases_Nivel3\\Files\\Libro1.xlsx")


ac = archivo["Hoja1"]

filas = ac.max_row
columnas = ac.max_column

print('Maximo de filas: ' + str(filas))
print('Maximo de columnas: ' + str(columnas))

#for r in ac['A1':'C4']:
    #for c in r:
        #print(c.value)

for r in range(1,filas+1):
    for c in range(1, columnas+1):
        v = ac.cell(r, c)
        print(v.value)



